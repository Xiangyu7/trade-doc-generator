# -*- coding: utf-8 -*-
"""
外贸单据自动生成 —— 第1步：PI / CI / PL
================================================
读取「订单信息表.xlsx」里的每一票订单，套用主文件夹里的「CPH模板.xlsx」，
自动生成形式发票(PI)、商业发票(CI)、装箱单(PL)，并导出成 PDF。

用法：
    在「单据自动生成」文件夹里打开终端，运行：
        python3 生成单据.py
    生成结果在「输出单据」文件夹里，每票订单一个 Excel + 一个 PDF。

原理：
    CPH模板里「发票B」和「箱单B」的字段大多是公式引用「形式发票」，
    所以脚本只填「形式发票」这一张主表，CI/PL 会自动联动；
    箱单里独立的「体积CBM」单独填。
"""
import sys
import shutil
import subprocess
from pathlib import Path

import openpyxl

# ---------------- 路径配置 ----------------
BASE = Path(__file__).resolve().parent
TEMPLATE = BASE.parent / "CPH模板.xlsx"        # 主文件夹里的模板
ORDERS = BASE / "订单信息表.xlsx"
OUTDIR = BASE / "输出单据"
SOFFICE = "/Applications/LibreOffice.app/Contents/MacOS/soffice"

# 订单表「列名」 -> (形式发票主表的目标单元格, 格式模板)
# 形式发票是主表，CI/PL 自动联动，无需单独填
FIELD_MAP = {
    "发票号":         ("A4",  "Invoice No.:{}"),
    "发票日期":       ("C4",  "PI Date.:{}"),
    "ReferenceNo":   ("E4",  "Reference No.:{}"),
    "买方名称及地址":  ("D6",  "{}"),
    "通知方":         ("A14", "{}"),
    "品名及HS编码":   ("A20", "{}"),
    "数量KGS":        ("D20", "{}"),
    "单价USD":        ("E20", "{}"),
    "付款条款":       ("B25", "{}"),
    "贸易术语":       ("B26", "{}"),
    "包装方式":       ("B27", "{}"),
    "装货港":         ("B28", "{}"),
    "卸货港":         ("B29", "{}"),
    "唛头":           ("B30", "{}"),
}
# 「箱单 B」上需要单独填的（不联动主表）
PL_MAP = {
    "体积CBM": ("F20", "{}"),
}
# 这些字段写成数字，其它写成文本
NUMERIC = {"数量KGS", "单价USD", "体积CBM"}


def safe_name(s: str) -> str:
    """把发票号变成安全的文件名。"""
    bad = '/\\:*?"<>|'
    return "".join("_" if ch in bad else ch for ch in str(s)).strip() or "无发票号"


def read_orders():
    """读订单表，返回 [dict, ...]，按表头取值。"""
    wb = openpyxl.load_workbook(ORDERS, data_only=True)
    ws = wb.active
    headers = [ (c.value.strip() if isinstance(c.value, str) else c.value)
                for c in ws[1] ]
    orders = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v in (None, "") for v in row):
            continue
        orders.append(dict(zip(headers, row)))
    return orders


def fill_sheet(ws, order, field_map):
    """按映射把订单数据填进某张 sheet 的主表单元格。"""
    for col, (cell, fmt) in field_map.items():
        val = order.get(col)
        if val in (None, ""):
            continue
        if col in NUMERIC:
            try:
                num = float(val)
                ws[cell] = int(num) if num.is_integer() else num
            except (TypeError, ValueError):
                ws[cell] = val
        else:
            ws[cell] = fmt.format(val)


def set_print(ws):
    """设定打印区域 + 适应页宽，PDF 才干净。"""
    ws.print_area = "A1:F40"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)


def to_pdf(xlsx_path: Path) -> bool:
    """用 LibreOffice 把 xlsx 转 pdf（一个PDF含PI/CI/PL三页）。"""
    if not Path(SOFFICE).exists():
        return False
    try:
        subprocess.run(
            [SOFFICE, "--headless", "--calc", "--convert-to", "pdf",
             "--outdir", str(OUTDIR), str(xlsx_path)],
            check=True, capture_output=True, timeout=120,
        )
        return (OUTDIR / (xlsx_path.stem + ".pdf")).exists()
    except Exception as e:
        print(f"   ⚠️  转PDF失败：{e}")
        return False


def main():
    print("=" * 50)
    print(" 外贸单据自动生成 —— PI / CI / PL")
    print("=" * 50)

    for p, label in [(TEMPLATE, "CPH模板.xlsx"), (ORDERS, "订单信息表.xlsx")]:
        if not p.exists():
            print(f"❌ 找不到 {label}：{p}")
            sys.exit(1)

    OUTDIR.mkdir(exist_ok=True)
    orders = read_orders()
    print(f"读到 {len(orders)} 行订单。\n")

    done = 0
    for i, order in enumerate(orders, start=2):  # 行号从2起（1是表头）
        inv = order.get("发票号")
        if inv in (None, ""):
            print(f"· 第{i}行：无发票号，跳过（预填行，补上发票号后再生成）")
            continue

        inv = str(inv).strip()
        print(f"▶ 第{i}行：发票号 {inv} 生成中…")

        wb = openpyxl.load_workbook(TEMPLATE)
        fill_sheet(wb["形式发票"], order, FIELD_MAP)   # 主表，CI/PL自动联动
        fill_sheet(wb["箱单 B"], order, PL_MAP)        # 箱单独立字段

        invoice_sheets = ("形式发票", "发票B", "箱单 B")
        for sn in invoice_sheets:
            set_print(wb[sn])
        # 其余表（DOC注意事项 / Sheet3）藏起来：保留在Excel里，但不进PDF
        for sn in wb.sheetnames:
            if sn not in invoice_sheets:
                wb[sn].sheet_state = "hidden"
        wb.active = wb.sheetnames.index("形式发票")

        xlsx_out = OUTDIR / f"{safe_name(inv)}_PI-CI-PL.xlsx"
        wb.save(xlsx_out)
        print(f"   ✓ Excel: {xlsx_out.name}")

        if to_pdf(xlsx_out):
            print(f"   ✓ PDF : {xlsx_out.stem}.pdf")
        done += 1

    print(f"\n完成：生成 {done} 票单据，在「{OUTDIR.name}」文件夹里。")


if __name__ == "__main__":
    main()
